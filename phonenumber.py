
from phone import Phone
from openpyxl import Workbook


from openpyxl import load_workbook  #导入模块
wb = load_workbook(filename = 'phone.xlsx') #打开phone.xlsx文件,默认可读写，若有需要可以指定write_only和read_only为True
sheet = wb['Sheet1'] #找到工作表



for n in range(1,sheet.max_row+1):
    number = sheet.cell(row=n,column=1).value

    data = Phone().find(number)
    if data == None:
        continue
    sheet.cell(row=n, column=2).value = data['city']
    sheet.cell(row=n, column=3).value = data['province']
    sheet.cell(row=n, column=4).value = data['phone_type']
# 保存在result.xlsx文件中
wb.save('result.xlsx')
